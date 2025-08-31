"""
mcp_excel_toolkit.py
FastMCP server – Excel/CSV extraction → Markdown (no camel dependency)
"""

# --------------------------------------------------------------------------- #
#  Imports
# --------------------------------------------------------------------------- #
import os
from typing import Optional, Tuple

import anyio
import pandas as pd
from loguru import logger
from mcp.server.fastmcp import FastMCP


# --------------------------------------------------------------------------- #
#  Helper class
# --------------------------------------------------------------------------- #
class ExcelToolkit:
    """
    Extracts rich information from Excel (.xls/.xlsx) or CSV files:
      • Every sheet converted to Markdown using `tabulate`
      • List of all cell coordinates with value + font / fill RGB colours
    """

    def __init__(self, timeout: Optional[float] = None):
        self.timeout = timeout

    # ---------- public synchronous API ------------------------------------ #
    def extract_excel_content(self, document_path: str) -> str:
        if not document_path.lower().endswith((".xls", ".xlsx", ".csv")):
            raise ValueError("Only .xls, .xlsx or .csv files are supported.")

        logger.info(f"Processing Excel/CSV file: {document_path}")

        if document_path.lower().endswith(".csv"):
            return self._handle_csv(document_path)

        # If it's .xls, convert to .xlsx first
        if document_path.lower().endswith(".xls"):
            from xls2xlsx import XLS2XLSX

            out_path = document_path.rsplit(".", 1)[0] + ".xlsx"
            XLS2XLSX(document_path).to_xlsx(out_path)
            document_path = out_path
            logger.debug(f"Converted .xls → .xlsx : {out_path}")

        return self._handle_xlsx(document_path)

    # ---------- helpers ---------------------------------------------------- #
    def _handle_csv(self, path: str) -> str:
        try:
            df = pd.read_csv(path)
        except Exception as e:
            logger.error(f"CSV read failed: {e}")
            raise

        return "CSV File Processed:\n" + self._df_to_md(df)

    def _handle_xlsx(self, path: str) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        output_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cells_info = []

            for row in ws.iter_rows():
                for cell in row:
                    coord = f"{cell.row}{cell.column_letter}"
                    font_rgb = (
                        cell.font.color.rgb
                        if cell.font and cell.font.color and cell.font.color.rgb
                        else None
                    )
                    fill_rgb = (
                        cell.fill.fgColor.rgb
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb
                        else None
                    )
                    cells_info.append(
                        {
                            "index": coord,
                            "value": cell.value,
                            "font_color": font_rgb,
                            "fill_color": fill_rgb,
                        }
                    )

            # Re‑load sheet via pandas for prettier Markdown
            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")

            part = (
                f"Sheet Name: {sheet_name}\n"
                f"Cell information list:\n{cells_info}\n\n"
                f"Markdown View of the content:\n{self._df_to_md(df)}\n"
                + "-" * 40
            )
            output_parts.append(part)

        return "\n".join(output_parts)

    @staticmethod
    def _df_to_md(df: pd.DataFrame) -> str:
        from tabulate import tabulate

        return tabulate(df, headers="keys", tablefmt="pipe")


# --------------------------------------------------------------------------- #
#  FastMCP server
# --------------------------------------------------------------------------- #
mcp = FastMCP("excel_toolkit")
toolkit = ExcelToolkit()


@mcp.tool()
async def extract_excel_content(document_path: str) -> str:
    """
    Return a Markdown‑rich description of *document_path* (xls/xlsx/csv).
    """
    # Run the synchronous extractor in a worker thread so the event loop
    # stays free.
    try:
        return await anyio.to_thread.run_sync(
            toolkit.extract_excel_content, document_path
        )
    except Exception as e:
        raise ValueError(str(e))


# --------------------------------------------------------------------------- #
#  Entrypoint
# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    mcp.run(transport="stdio")
